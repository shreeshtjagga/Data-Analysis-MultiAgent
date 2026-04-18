

import io
import asyncio
import logging
import os
import json
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Annotated

import pandas as pd
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile, status, Request, Response, Query, BackgroundTasks
from fastapi.responses import JSONResponse
from fastapi.middleware.gzip import GZipMiddleware
import traceback
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from sqlalchemy.ext.asyncio import AsyncSession

from .core import cache as redis_cache
from .analysis_history import (
    _serialize_charts,
    compute_file_hash,
    delete_analysis,
    get_analysis_by_id,
    get_analysis_by_hash,
    get_user_analysis_history,
    save_analysis,
)
from .auth import (
    create_access_token,
    get_user_by_id,
    login_user,
    register_user,
    request_password_reset,
    reset_password_with_token,
    verify_access_token,
    verify_google_token,
    login_google_user,
    create_refresh_token,
    verify_refresh_token,
    REFRESH_EXPIRE_DAYS,
)
from .core.constants import APP_VERSION, PIPELINE_VERSION
from .core.graph import run_pipeline
from .core.logging_config import configure_logging
from .core.upload_parsing import read_csv_with_fallback, validate_upload_magic
from .core.utils import sanitize_floats, truncate_stats_for_llm
from .core.data_agent import run_data_query
from .db import get_db, init_db
from .models.schemas import (
    AnalysisListResponse,
    AuthResponse,
    ChatRequest,
    DeleteResponse,
    ForgotPasswordRequest,
    ForgotPasswordResponse,
    GoogleLoginRequest,
    HealthResponse,
    ResetPasswordRequest,
    TokenResponse,
    UserLogin,
    UserRegister,
    UserResponse,
)

configure_logging()
logger = logging.getLogger(__name__)

APP_ENV = os.getenv("APP_ENV", "production")
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "").strip()
FRONTEND_GOOGLE_CLIENT_ID = (
    os.getenv("FRONTEND_GOOGLE_CLIENT_ID", "").strip()
    or os.getenv("VITE_GOOGLE_CLIENT_ID", "").strip()
)
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(10 * 1024 * 1024))) # 10 MB limit for strict RAM bounds
MAX_ANALYZE_ROWS = int(os.getenv("MAX_ANALYZE_ROWS", "15000")) # Lowered to 15K rows for 500MB Render memory limit
MAX_ANALYZE_COLUMNS = int(os.getenv("MAX_ANALYZE_COLUMNS", "150"))
MAX_EXCEL_SHEETS = int(os.getenv("MAX_EXCEL_SHEETS", "5"))
MAX_QUESTION_CHARS = int(os.getenv("CHAT_MAX_QUESTION_CHARS", "1200"))
MAX_CONTEXT_BYTES = int(os.getenv("CHAT_MAX_CONTEXT_BYTES", str(2 * 1024 * 1024)))  # 2 MB context limit
READ_CHUNK_BYTES = 1024 * 1024
CHAT_RATE_LIMIT = int(os.getenv("CHAT_RATE_LIMIT", "10"))
CHAT_RATE_WINDOW = int(os.getenv("CHAT_RATE_WINDOW_SECONDS", "60"))

# Groq client is stateless — create once at startup, reuse across requests
_groq_client = None

def _get_groq_client():
    global _groq_client
    if _groq_client is None:
        api_key = os.getenv("GROQ_API_KEY")
        if api_key:
            from groq import Groq
            _groq_client = Groq(api_key=api_key)
    return _groq_client


_raw_origins = os.getenv("CORS_ORIGINS", "http://localhost:5173,http://localhost:3000")
origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]


@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("Starting DataPulse API v2 (pipeline %s)", PIPELINE_VERSION)

    if not os.getenv("GROQ_API_KEY"):
        raise RuntimeError(
            "GROQ_API_KEY environment variable is not set. "
            "The application cannot start without it. "
            "Set it in your .env file or environment."
        )

    if APP_ENV == "production" and "*" in origins:
        raise RuntimeError("CORS_ORIGINS cannot contain '*' in production")

    if GOOGLE_CLIENT_ID and FRONTEND_GOOGLE_CLIENT_ID and GOOGLE_CLIENT_ID != FRONTEND_GOOGLE_CLIENT_ID:
        raise RuntimeError(
            "Google OAuth Client ID mismatch between backend and frontend configuration"
        )

    # ── Eagerly warm up connections so first request is instant ──────────────
    await init_db()

    # Warm DB pool: open one real connection now so asyncpg doesn't cold-start
    try:
        from sqlalchemy import text
        from .db import engine
        async with engine.connect() as conn:
            await conn.execute(text("SELECT 1"))
        logger.info("DB connection pool warmed up")
    except Exception as exc:
        logger.warning("DB warmup failed (non-fatal): %s", exc)

    # Warm Redis: open the connection now instead of on first request
    try:
        redis_ok = await redis_cache.ping()
        logger.info("Redis warmed up (reachable=%s)", redis_ok)
    except Exception as exc:
        logger.warning("Redis warmup failed (non-fatal): %s", exc)
    # ─────────────────────────────────────────────────────────────────────────

    yield
    await redis_cache.close()
    logger.info("DataPulse API shutdown complete")




app = FastAPI(
    title="DataPulse API",
    description="Multi-agent CSV analysis API",
    version=APP_VERSION,
    lifespan=lifespan,
)

app.add_middleware(GZipMiddleware, minimum_size=1024)  # compress responses > 1KB
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    return response

@app.exception_handler(Exception)
async def catch_all_exception_handler(request: Request, exc: Exception):
    error_trace = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    logger.error(f"Unhandled error: {error_trace}")
    
    if os.getenv("APP_ENV", "production") == "development":
        return JSONResponse(status_code=500, content={"detail": str(exc), "trace": error_trace})
        
    return JSONResponse(status_code=500, content={"detail": "Internal Server Error"})



security = HTTPBearer()


async def get_current_user_id(
    credentials: Annotated[HTTPAuthorizationCredentials, Depends(security)],
) -> int:
    token = credentials.credentials
    payload = verify_access_token(token)
    if payload is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return int(payload["sub"])


async def check_ip_rate_limit(request: Request):
    client_ip = request.client.host if request.client else "unknown"
    key = f"ratelimit:ip:{client_ip}"
    try:
        count = await redis_cache.increment_with_ttl(key, 60)
        if count > 5:
            raise HTTPException(status_code=429, detail="Too many requests from this IP. Please try again in a minute.")
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise exc
        logger.warning(f"Rate limiting failed for {key}: {exc}")

async def check_user_rate_limit(user_id: int = Depends(get_current_user_id)):
    key = f"ratelimit:user_analyze:{user_id}"
    try:
        count = await redis_cache.increment_with_ttl(key, 60)
        if count > 5:
            raise HTTPException(status_code=429, detail="Too many analysis requests. Please try again in a minute.")
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise exc
        logger.warning(f"Rate limiting failed for {key}: {exc}")





@app.get("/health", response_model=HealthResponse, tags=["system"])
async def health_check(db: AsyncSession = Depends(get_db)):
    """Liveness probe — returns Postgres and Redis reachability."""
    pg_ok = False
    try:
        await db.execute(__import__("sqlalchemy").text("SELECT 1"))
        pg_ok = True
    except Exception:
        pass

    redis_ok = await redis_cache.ping()

    return HealthResponse(
        status="ok" if (pg_ok and redis_ok) else "degraded",
        postgres=pg_ok,
        redis=redis_ok,
    )




@app.post("/auth/register", response_model=AuthResponse, status_code=status.HTTP_201_CREATED, tags=["auth"], dependencies=[Depends(check_ip_rate_limit)])
async def register(body: UserRegister, db: AsyncSession = Depends(get_db)):
    result = await register_user(db, body.email, body.password, body.name)
    if not result["success"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result["message"])
    
    # Build the user response object
    now = datetime.utcnow()
    user_response = UserResponse(
        id=result["user_id"],
        name=result.get("name"),
        email=result["email"],
        created_at=now,
        updated_at=now
    )
    
    return AuthResponse(
        success=True, 
        message=result["message"],
        user=user_response
    )


@app.post("/auth/login", response_model=TokenResponse, tags=["auth"], dependencies=[Depends(check_ip_rate_limit)])
async def login(body: UserLogin, db: AsyncSession = Depends(get_db), response: Response = None):
    result = await login_user(db, body.email, body.password)
    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=result["message"],
            headers={"WWW-Authenticate": "Bearer"},
        )
    user_obj = result["user"]
    # create an HttpOnly refresh cookie and return the short-lived access token
    try:
        refresh_token = create_refresh_token(user_obj["id"], user_obj["email"])

        if response is not None:
            response.set_cookie(
                key="datapulse_refresh",
                value=refresh_token,
                httponly=True,
                secure=(APP_ENV == "production"),
                samesite="lax",
                max_age=REFRESH_EXPIRE_DAYS * 24 * 3600,
                path="/",
            )
    except Exception:
        logger.exception("Failed to create refresh token")
    return TokenResponse(
        access_token=result["access_token"],
        token_type="bearer",
        user=UserResponse(
            id=user_obj["id"],
            name=user_obj.get("name"),
            email=user_obj["email"],
            created_at=user_obj["created_at"],
            updated_at=user_obj["updated_at"],
        ),
    )


@app.post("/auth/forgot-password", response_model=ForgotPasswordResponse, tags=["auth"], dependencies=[Depends(check_ip_rate_limit)])
async def forgot_password(body: ForgotPasswordRequest, db: AsyncSession = Depends(get_db)):
    result = await request_password_reset(db, body.email)
    return ForgotPasswordResponse(
        success=True,
        message=result.get("message", "If an account exists for that email, a password reset link has been sent."),
        debug_reset_token=result.get("debug_reset_token"),
    )


@app.post("/auth/reset-password", response_model=AuthResponse, tags=["auth"], dependencies=[Depends(check_ip_rate_limit)])
async def reset_password(body: ResetPasswordRequest, db: AsyncSession = Depends(get_db)):
    result = await reset_password_with_token(db, body.token, body.new_password)
    if not result.get("success"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.get("message", "Password reset failed"))
    return AuthResponse(success=True, message=result["message"])


@app.post("/auth/google", response_model=TokenResponse, tags=["auth"])
async def login_with_google(body: GoogleLoginRequest, db: AsyncSession = Depends(get_db), response: Response = None):
    credential = body.credential
    if not credential:
        raise HTTPException(status_code=400, detail="Missing Google credential")

    frontend_client_id = (body.client_id or "").strip()
    if frontend_client_id and GOOGLE_CLIENT_ID and frontend_client_id != GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=400, detail="Google client ID mismatch")

    idinfo = verify_google_token(credential)
    if not idinfo:
        raise HTTPException(status_code=401, detail="Invalid Google token")
    if GOOGLE_CLIENT_ID and idinfo.get("aud") != GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=401, detail="Invalid Google token audience")

    email = idinfo.get("email")
    google_id = idinfo.get("sub")
    name = idinfo.get("name")
    if not email:
        raise HTTPException(status_code=400, detail="No email provided by Google")

    result = await login_google_user(db, email, google_id, name)
    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail=result["message"],
        )
    
    user_obj = result["user"]
    try:
        refresh_token = create_refresh_token(user_obj["id"], user_obj["email"])
        if response is not None:
            response.set_cookie(
                key="datapulse_refresh",
                value=refresh_token,
                httponly=True,
                secure=(APP_ENV == "production"),
                samesite="lax",
                max_age=REFRESH_EXPIRE_DAYS * 24 * 3600,
                path="/",
            )
    except Exception:
        logger.exception("Failed to create refresh token for google login")

    return TokenResponse(
        access_token=result["access_token"],
        token_type="bearer",
        user=UserResponse(
            id=user_obj["id"],
            name=user_obj.get("name"),
            email=user_obj["email"],
            created_at=user_obj["created_at"],
            updated_at=user_obj["updated_at"],
        ),
    )



@app.post("/auth/refresh", response_model=TokenResponse, tags=["auth"])
async def refresh_token(request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    token = request.cookies.get("datapulse_refresh")
    if not token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing refresh token")
    payload = verify_refresh_token(token)
    if payload is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired refresh token")
    user_id = int(payload["sub"])
    user = await get_user_by_id(db, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    access_token = create_access_token(user.id, user.email)
    try:
        new_refresh = create_refresh_token(user.id, user.email)
        response.set_cookie(
            key="datapulse_refresh",
            value=new_refresh,
            httponly=True,
            secure=(APP_ENV == "production"),
            samesite="lax",
            max_age=REFRESH_EXPIRE_DAYS * 24 * 3600,
            path="/",
        )
    except Exception:
        logger.exception("Failed to rotate refresh token")

    return TokenResponse(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(
            id=user.id,
            name=user.name,
            email=user.email,
            created_at=user.created_at,
            updated_at=user.updated_at,
        ),
    )


@app.post("/auth/logout", tags=["auth"])
async def logout(response: Response):

    response.delete_cookie("datapulse_refresh", path="/")
    return {"success": True, "message": "Logged out"}


@app.get("/auth/me", response_model=UserResponse, tags=["auth"])
async def me(
    user_id: Annotated[int, Depends(get_current_user_id)],
    db: AsyncSession = Depends(get_db),
):
    user = await get_user_by_id(db, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    return UserResponse(
        id=user.id,
        name=user.name,
        email=user.email,
        created_at=user.created_at,
        updated_at=user.updated_at,
    )





# ── Utility for background storage ───────────────────────────────────────
# Use an absolute path so the file is always written to the same location
# that data_agent.py reads from, regardless of the server's working directory.
_API_FILE_DIR = os.path.dirname(os.path.abspath(__file__))   # .../backend/
_PARQUET_STORAGE_DIR = os.path.join(_API_FILE_DIR, "storage", "data")

def persist_full_data_backend(df: pd.DataFrame, file_hash: str):
    """Saves cleaned DataFrame to Parquet in the background."""
    try:
        os.makedirs(_PARQUET_STORAGE_DIR, exist_ok=True)
        storage_path = os.path.join(_PARQUET_STORAGE_DIR, f"{file_hash}.parquet")
        df.to_parquet(storage_path, index=False)
        logger.info("Background storage: Saved full data to %s", storage_path)
    except Exception as exc:
        logger.warning("Background storage failed for %s: %s", file_hash, exc)


@app.post("/analyze", tags=["analysis"], dependencies=[Depends(check_user_rate_limit)])
async def analyze(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):

    filename = file.filename or ""
    parsed_ext = filename.lower().split('.')[-1] if '.' in filename else ""
    if parsed_ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are accepted")

    total = 0
    chunks: list[bytes] = []
    while True:
        chunk = await file.read(READ_CHUNK_BYTES)
        if not chunk:
            break
        total += len(chunk)
        if total > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Max allowed size is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB",
            )
        chunks.append(chunk)
    file_bytes = b"".join(chunks)

    if len(file_bytes) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    validate_upload_magic(parsed_ext, file_bytes)

    file_hash = compute_file_hash(file_bytes, filename)


    cached = await get_analysis_by_hash(db, user_id, file_hash)
    if (
        cached
        and cached.get("stats_summary")
        and cached.get("insights")
        and not (cached.get("errors") or [])
        and cached.get("pipeline_version") == PIPELINE_VERSION
    ):
        cached["charts"] = {k: v for k, v in (cached.get("charts") or {}).items()}
        return {"from_cache": True, **cached}
    elif cached:
        logger.info("Cache not eligible for reuse for user %d / %s — re-running pipeline", user_id, file.filename)

    await db.rollback()


    try:
        if parsed_ext == "csv":
            df = read_csv_with_fallback(
                file_bytes,
                max_analyze_rows=MAX_ANALYZE_ROWS,
                max_analyze_columns=MAX_ANALYZE_COLUMNS,
            )
        else:
            if parsed_ext == "xlsx":
                from openpyxl import load_workbook

                wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                try:
                    sheet_count = len(wb.sheetnames)
                    if sheet_count > MAX_EXCEL_SHEETS:
                        raise HTTPException(
                            status_code=413,
                            detail=f"Excel file has {sheet_count} sheets. Maximum allowed is {MAX_EXCEL_SHEETS}.",
                        )

                    active_sheet = wb[wb.sheetnames[0]]
                    if (active_sheet.max_row or 0) > MAX_ANALYZE_ROWS:
                        raise HTTPException(
                            status_code=413,
                            detail=f"Excel file has {active_sheet.max_row} rows. Maximum allowed is {MAX_ANALYZE_ROWS}.",
                        )
                    if (active_sheet.max_column or 0) > MAX_ANALYZE_COLUMNS:
                        raise HTTPException(
                            status_code=413,
                            detail=f"Excel file has {active_sheet.max_column} columns. Maximum allowed is {MAX_ANALYZE_COLUMNS}.",
                        )
                finally:
                    wb.close()

                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    engine="openpyxl",
                    sheet_name=0,
                    nrows=MAX_ANALYZE_ROWS + 1,
                )
            else:
                try:
                    import xlrd  # noqa: F401
                except Exception:
                    raise HTTPException(
                        status_code=422,
                        detail="Legacy .xls upload requires the 'xlrd' package. Please upload CSV/XLSX or install xlrd on the backend.",
                    )

                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    engine="xlrd",
                    sheet_name=0,
                    nrows=MAX_ANALYZE_ROWS + 1,
                )
    except Exception as exc:
        if isinstance(exc, HTTPException):
            raise exc
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")

    row_count, column_count = df.shape
    if row_count > MAX_ANALYZE_ROWS:
        raise HTTPException(
            status_code=413,
            detail=f"Dataset has {row_count} rows. Maximum allowed is {MAX_ANALYZE_ROWS}.",
        )
    if column_count > MAX_ANALYZE_COLUMNS:
        raise HTTPException(
            status_code=413,
            detail=f"Dataset has {column_count} columns. Maximum allowed is {MAX_ANALYZE_COLUMNS}.",
        )

    # Run the agentic pipeline
    import asyncio
    state = await asyncio.to_thread(run_pipeline, df)

    # ── Full Data Persistence (Background) ───────────────────────────────────
    if getattr(state, "clean_df", None) is not None:
        background_tasks.add_task(persist_full_data_backend, state.clean_df.copy(), file_hash)

    # ── Strip full DataFrames BEFORE model_dump to prevent serialising
    # 30k rows as Python dicts (≈500 MB RAM spike). ───────────────────────────
    preview_raw = json.loads(
        state.raw_df.head(100).to_json(orient="records")
    ) if getattr(state, "raw_df", None) is not None else []
    
    preview_clean = json.loads(
        state.clean_df.head(100).to_json(orient="records")
    ) if getattr(state, "clean_df", None) is not None else []
    
    state.raw_df   = None
    state.clean_df = None

    result = state.model_dump()
    result["raw_df"]   = preview_raw
    result["clean_df"] = preview_clean
    result["file_hash"] = file_hash  # Crucial for chat context

    # Only hard-fail if BOTH stats and insights are completely empty.
    # Partial data (e.g., architect succeeded but statistician failed) still yields a usable page.
    has_stats    = bool(result.get("stats_summary") and result["stats_summary"].get("row_count"))
    has_insights = bool(result.get("insights") and result["insights"].get("findings"))
    # True last-resort: literally nothing was produced
    is_fatal = not has_stats and not has_insights

    if is_fatal:
        logger.error("Pipeline critically failed for user %d / %s: %s", user_id, filename, state.errors)
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Analysis pipeline failed — no statistics or insights were produced. Please check the dataset format and try again.",
                "errors": [str(e) for e in state.errors],
            },
        )
    elif state.errors:
        # Non-fatal warnings: log them but continue — the result is still usable
        logger.warning(
            "Pipeline completed with non-fatal errors for user %d / %s: %s",
            user_id, filename, state.errors,
        )

    # Fix #6: compute serialized charts once, reuse for both DB save and response
    serialized_charts = _serialize_charts(result.get("charts") or {})

    save_result = await save_analysis(
        db=db,
        user_id=user_id,
        file_name=filename,
        file_hash=file_hash,
        file_size=len(file_bytes),
        analysis_result=result,
        serialized_charts=serialized_charts,
    )
    if not save_result["success"]:
        logger.warning("Failed to persist analysis: %s", save_result["message"])
    else:
        result["analysis_id"] = save_result.get("analysis_id")

    result["charts"] = serialized_charts
    result["partial"] = False
    if state.errors:
        result["warnings"] = state.errors   # surface non-fatal errors to frontend

    # orjson serialises NaN/Inf → null natively and is ~10x faster than
    # the manual recursive sanitize_floats walk on large result dicts.
    try:
        import orjson
        safe_result = orjson.loads(orjson.dumps(result, option=orjson.OPT_NON_STR_KEYS))
    except Exception:
        safe_result = sanitize_floats(result)   # fallback if orjson not available

    return {"from_cache": False, "pipeline_version": PIPELINE_VERSION, **safe_result}




@app.get("/history", response_model=AnalysisListResponse, tags=["history"])
async def history(
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
    limit: int = Query(default=20, ge=1, le=100),
):
    items = await get_user_analysis_history(db, user_id, limit=limit)
    return AnalysisListResponse(
        success=True,
        message="History retrieved",
        total=len(items),
        analyses=items,
    )


@app.get("/history/{analysis_id}", tags=["history"])
async def history_item(
    analysis_id: int,
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    item = await get_analysis_by_id(db, user_id, analysis_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Analysis not found or access denied")
    item["charts"] = {k: v for k, v in (item.get("charts") or {}).items()}
    return item


@app.delete("/history/{analysis_id}", response_model=DeleteResponse, tags=["history"])
async def remove_analysis(
    analysis_id: int,
    user_id: int = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
):
    result = await delete_analysis(db, user_id, analysis_id)
    if not result["success"]:
        raise HTTPException(status_code=404, detail=result["message"])
    return DeleteResponse(success=True, message=result["message"])


@app.post("/chat", tags=["analysis"])
async def chat_with_analysis(
    body: ChatRequest,
    user_id: int = Depends(get_current_user_id),
):
    """Answer a user question about the current analysis context."""

    try:
        key = f"ratelimit:chat:{user_id}"
        count = await redis_cache.increment_with_ttl(key, CHAT_RATE_WINDOW)
        if count > CHAT_RATE_LIMIT:
            raise HTTPException(status_code=429, detail=f"Too many chat requests. Please wait {CHAT_RATE_WINDOW} seconds.")
    except Exception as exc:
        # Graceful degradation: keep chat functional even if Redis is unavailable.
        logger.warning("Redis rate limiter unavailable for user %s, continuing without rate limit: %s", user_id, exc)

    question = body.question.strip()
    context = body.context or {}

    if not question:
        raise HTTPException(status_code=400, detail="Question is required")
    if len(question) > MAX_QUESTION_CHARS:
        raise HTTPException(
            status_code=413,
            detail=f"Question too long. Max length is {MAX_QUESTION_CHARS} characters",
        )

    context_blob = json.dumps(context, default=str)
    if len(context_blob.encode("utf-8")) > MAX_CONTEXT_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Context too large. Max size is {MAX_CONTEXT_BYTES // 1024} KB",
        )

    stats = context.get("stats") or context.get("stats_summary") or {}
    insights = context.get("insights") or {}
    file_name = context.get("fileName") or "dataset"


    slim_stats = truncate_stats_for_llm(stats)
    profile = slim_stats.get("dataset_profile") or {}
    outlier_counts = slim_stats.get("outlier_counts") or {}
    if not outlier_counts:
        outlier_counts = {
            k: int((v or {}).get("count", 0))
            for k, v in (stats.get("outliers") or {}).items()
        }
    top_outliers = sorted(outlier_counts.items(), key=lambda kv: kv[1], reverse=True)[:10]
    correlations = context.get("correlations") or slim_stats.get("strong_correlations") or []
    quality = context.get("dataQuality") or slim_stats.get("data_quality") or {}
    
    # Extract structural summaries of charts so the LLM understands exactly what data is inside them
    charts_data = context.get("charts", {})
    charts_summary = {}
    if isinstance(charts_data, dict):
        for k, v in charts_data.items():
            try:
                c = json.loads(v) if isinstance(v, str) else v
                details = []
                for trace in c.get("data", []):
                    ttype = trace.get("type", "unknown")
                    if ttype in ("pie", "funnelarea"):
                        labels = trace.get("labels") or trace.get("x") or []
                        values = trace.get("values") or trace.get("y") or []
                        if isinstance(labels, list) and isinstance(values, list):
                            pairs = [f"{l}: {val}" for l, val in zip(labels[:10], values[:10])]
                            details.append(f"Pie/Donut Breakdown: {', '.join(pairs)}")
                        else:
                            details.append("Pie/Donut chart")
                    elif ttype in ("bar", "scatter", "violin", "box"):
                        x = (trace.get("x") or [])[:8]
                        y = (trace.get("y") or [])[:8]
                        details.append(f"{ttype.capitalize()} Data: X={x}, Y={y}")
                    elif ttype in ("heatmap", "choropleth"):
                        z_sample = (trace.get("z") or [])[:3]
                        details.append(f"Heatmap with z-values (sample): {z_sample}")
                    elif ttype == "histogram":
                        x = (trace.get("x") or [])[:8]
                        details.append(f"Histogram of: {x}")
                    else:
                        details.append(f"{ttype} chart")
                # Safely extract title: layout.title can be a str or dict
                layout_title = c.get("layout", {}).get("title", {})
                if isinstance(layout_title, dict):
                    title = layout_title.get("text") or k
                elif isinstance(layout_title, str) and layout_title:
                    title = layout_title
                else:
                    title = k
                charts_summary[k] = f"[key='{k}'] Title '{title}' \u2014 " + (" | ".join(details) if details else "chart")
            except Exception as e:
                import traceback
                logger.warning("Chart summary parse failed for key '%s': %s", k, traceback.format_exc())
                charts_summary[k] = f"[key='{k}'] Visual chart (parse error)"
    else:
        charts_summary = charts_data

    # Build a concise list of EXACT chart keys so the LLM never invents one
    exact_chart_keys = list(charts_summary.keys()) if isinstance(charts_summary, dict) else []
    outlier_summary = context.get("outlierSummary") or [
        {"column": col, "count": count} for col, count in top_outliers
    ]

    # ── Helper: extract the most recently mentioned entity name from history ──
    def _extract_recent_entity(history: list) -> str | None:
        """Scan recent assistant/user messages for a proper noun that could be a subject."""
        import re
        # Walk history in reverse — the most recent mention wins
        for m in reversed(history or []):
            content = m.get("content", "")
            # Match capitalised words (potential names / entities) — avoid common English words
            _STOP = {"the", "a", "an", "is", "are", "was", "were", "has", "have", "had",
                     "in", "on", "at", "of", "for", "and", "or", "but", "with", "by",
                     "according", "to", "data", "dataset", "fetched", "per", "as"}
            tokens = re.findall(r"[A-Z][a-z]+(?:\s[A-Z][a-z]+)*", content)
            for tok in reversed(tokens):
                if tok.lower() not in _STOP and len(tok) > 2:
                    return tok
        return None

    # Enhanced System Prompt for strict grounding and context awareness
    system_prompt = (
        "You are an elite Senior Data Analyst. You answer questions strictly grounded in the provided dataset context.\n"
        "Tone: Professional, authoritative, yet accessible. "
        "Brevity: Keep responses strictly between 1-4 sentences.\n"
        "Grounding Rules:\n"
        "1. MANDATORY: If you receive an 'ADDITIONAL DATA FROM FULL DATASET QUERY' block, you MUST use "
        "   those exact values to answer the question. Do NOT say 'I don't have that information' — the data is right there.\n"
        "2. For top_n / bottom_n results: look at 'top_entry' → that is the single best/worst entity. "
        "   State its name and value explicitly. For 'ranked_results', list the top entries naturally.\n"
        "3. For group_aggregate results: 'result' is a dict of {entity: value}. Pick the highest/lowest as needed.\n"
        "4. For distinct / value_counts: summarize the list or count concisely.\n"
        "5. For search results: each item in 'result' is a full row dict — read the relevant field and state it directly.\n"
        "6. NEVER say 'I don't have that data' if an ADDITIONAL DATA block is present, even if the result looks unfamiliar. Parse it.\n"
        "7. If truly no data is available and it's not in stats, say: 'I couldn't find that in the dataset.'\n"
        "8. If the user is chatty (hi, thanks), be polite but do not spontaneously analyze data.\n"
        f"9. Referencing Charts: Use exactly `[CHART: key]` using only these keys: {exact_chart_keys}. Never modify them.\n"
        "10. Use previous messages to maintain continuity. Pronouns like 'his', 'her', 'their', 'its', 'he', 'she' "
        "    ALWAYS refer to the most recently mentioned person/entity in the conversation — look back and resolve them."
    )

    # Data Context Block (separated from the user's actual question to prevent prompt injection/confusion)
    data_context = (
        f"Context for file '{file_name}':\n"
        f"- Profile: {profile.get('label', 'unknown')} ({profile.get('domain', 'general')})\n"
        f"- Quality: {quality}\n"
        f"- Charts Summary: {charts_summary}\n"
        f"- Stats: {slim_stats}\n"
        f"- Insights: {insights}\n"
    )

    history_msgs = []
    if body.history:
        for m in body.history:
            # Validate roles to match standard LLM expectations
            role = "assistant" if m.get("role") in ["assistant", "ai"] else "user"
            history_msgs.append({"role": role, "content": m.get("content", "")})

    # Final message sequence
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "system", "content": f"DATA CONTEXT:\n{data_context}"}
    ]
    messages.extend(history_msgs)
    messages.append({"role": "user", "content": question})

    client = _get_groq_client()
    if client:
        try:
            # ── Phase 1: Intent & Data Retrieval ─────────────────────────────
            file_hash = context.get("file_hash")
            data_result = None

            # Optimization: Skip intent check for simple greetings
            is_greeting = any(g in question.lower() for g in ["hello", "hi", "hey", "thanks", "thank you"])

            if file_hash and not is_greeting:
                try:
                    # Build col_types from stats_summary sub-dicts
                    col_types = {}
                    for c in stats.get("numeric_columns", {}).keys():
                        col_types[c] = "numeric"
                    for c in stats.get("categorical_columns", {}).keys():
                        col_types[c] = "categorical"
                    for c in (stats.get("columns") or []):
                        if c not in col_types:
                            col_types[c] = "unknown"

                    # Build recent conversation context so follow-up questions resolve correctly
                    recent_history = ""
                    if body.history:
                        recent_turns = body.history[-8:]  # last 4 exchanges
                        recent_history = "\n".join(
                            f"{m.get('role','user').upper()}: {m.get('content','')}"
                            for m in recent_turns
                        )

                    # ── Pronoun detection: if question uses pronouns without a named entity,
                    #    inject the resolved subject into the question so the intent LLM has it.
                    import re as _re
                    _PRONOUNS = {"his", "her", "their", "its", "he", "she", "they",
                                 "him", "hers", "theirs", "this person", "that person"}
                    q_tokens = set(question.lower().split())
                    has_pronoun = bool(q_tokens & _PRONOUNS)
                    resolved_subject = None
                    if has_pronoun and body.history:
                        resolved_subject = _extract_recent_entity(body.history)
                    # Build the question the planner actually sees (with resolved subject if any)
                    planner_question = question
                    if resolved_subject and has_pronoun:
                        planner_question = f"{question} [Note: pronoun refers to '{resolved_subject}']"
                        logger.info("Pronoun resolved: '%s' → '%s'", question, resolved_subject)

                    intent_prompt = (
                        "You are a Data Query Planner. Analyze the user's question (considering conversation history for follow-ups) "
                        "and decide if querying the FULL dataset is needed.\n"
                        "If YES, return ONLY a single valid JSON object. If NO, return exactly 'NONE'.\n\n"
                        "CRITICAL RULES (read before deciding):\n"
                        "- ALWAYS query the dataset for attribute/biographical questions: age, birthday, date of birth, address, score, "
                        "  rank, nationality, team, role, salary, height, weight, position, stats of a named person/item.\n"
                        "- ALWAYS query for follow-up questions about a specific entity even if the name comes from pronouns (his/her/their).\n"
                        "- When a question uses pronouns (his/her/their/its/he/she), the subject is resolved in the [Note] annotation — USE IT.\n"
                        "- Prefer 'search' when looking up a named entity's full row (birthday, age, address, stats).\n"
                        "- Prefer 'filter_lookup' when mapping one specific column's value to another (e.g. get birthdate of a named player).\n"
                        "- NEVER return NONE for questions about a specific named person or item's attributes.\n\n"
                        "SUPPORTED QUERY TYPES — pick the BEST one:\n"
                        "0. filter_lookup  — look up one column's value by matching another. Use for: 'what is X for Y', 'find Z of W', "
                        "   'when is [person]'s birthday', 'what is [person]'s age/score/team'.\n"
                        "   NEVER use aggregate on ID/code columns — always use filter_lookup instead.\n"
                        '   Example: {"type":"filter_lookup","params":{"filter_col":"player_name","filter_val":"Lokesh Rahul","result_col":"date_of_birth"}}\n'
                        "1. top_n          — highest N rows by a numeric column. Use for: 'most expensive', 'highest', 'costliest', 'largest', 'best', 'maximum'.\n"
                        '   Example: {"type":"top_n","params":{"column":"Price","n":1}}   ← use n=1 for singular\n'
                        "2. bottom_n       — lowest N rows by a numeric column. Use for: 'cheapest', 'lowest', 'worst', 'least', 'minimum'.\n"
                        '   Example: {"type":"bottom_n","params":{"column":"Price","n":1}}\n'
                        "3. group_aggregate — group by a column, aggregate a numeric one. Use for: 'average price per brand', 'total sales per region'.\n"
                        '   Example: {"type":"group_aggregate","params":{"group_by":"Brand","column":"Revenue","func":"sum","n":10}}\n'
                        "   func options: sum, mean, max, min, count, median, std\n"
                        "4. filter_group   — filter rows then group+aggregate. Use for: 'top brand in 2021', 'most X in category Y'.\n"
                        '   Example: {"type":"filter_group","params":{"group_by":"Brand","func":"count","n":5,"filters":[{"column":"Year","op":"year","value":2021}]}}\n'
                        "5. value_counts   — count each unique category. Use for: 'how many of each X', 'distribution of Y', 'how many unique Z'.\n"
                        '   Example: {"type":"value_counts","params":{"column":"Category","n":15}}\n'
                        "6. aggregate      — single stat (no grouping). Use for: 'average price', 'total revenue', 'max score'. NEVER for IDs.\n"
                        '   Example: {"type":"aggregate","params":{"column":"Price","func":"mean"}}\n'
                        "7. distinct       — list unique values. Use for: 'what are all the brands', 'list all categories', 'what values does X have'.\n"
                        '   Example: {"type":"distinct","params":{"column":"Brand"}}\n'
                        "8. row_count      — count rows matching a condition.\n"
                        '   Example: {"type":"row_count","params":{"filters":[{"column":"Status","op":"eq","value":"Active"}]}}\n'
                        "9. search         — text search across all columns. Use for: 'find rows with X', 'show entries containing Y', "
                        "   'look up person Z', or whenever you need all fields of a matching row.\n"
                        '   Example: {"type":"search","params":{"value":"Lokesh Rahul","n":3}}\n'
                        "10. lookup        — fetch a specific row by index.\n"
                        '    Example: {"type":"lookup","params":{"row_index":0}}\n'
                        "11. correlation   — correlation between two numeric columns.\n"
                        '    Example: {"type":"correlation","params":{"column":"Price","column2":"Mileage"}}\n'
                        "12. percentile    — compute a percentile of a column.\n"
                        '    Example: {"type":"percentile","params":{"column":"Price","percentile":90}}\n\n'
                        "Filter ops: 'eq' (exact), 'neq' (not equal), 'contains' (text), 'gt/lt/gte/lte' (numeric), 'year', 'month', 'isnull', 'notnull'.\n\n"
                        "RULES:\n"
                        "- For superlatives ('costliest','cheapest','fastest','most expensive','lowest rated') → use top_n or bottom_n with n=1\n"
                        "- For 'which X has the most/highest/best Y' → use group_aggregate with func=max or sum\n"
                        "- For follow-up questions with pronouns → the [Note] in the question tells you who the subject is — USE IT\n"
                        "- For 'how many unique X' → use value_counts or distinct (use distinct when listing is needed, value_counts for counts)\n"
                        "- Column names in params need NOT be exact — the system will fuzzy-match them automatically\n"
                        "- Always use n=1 for singular questions like 'which ONE is...', 'what is THE most...'\n\n"
                        f"Dataset columns and types: {col_types}\n\n"
                        + (f"Recent conversation:\n{recent_history}\n\n" if recent_history else "")
                        + f"Current user question: {planner_question}\n\n"
                        "Return ONLY the JSON object or 'NONE'. No explanation, no markdown, no code fences."
                    )
                    # Use a smarter model for intent classification
                    intent_model = os.getenv("GROQ_INTENT_MODEL", "llama-3.3-70b-versatile")
                    intent_resp = client.chat.completions.create(
                        model=intent_model,
                        messages=[{"role": "user", "content": intent_prompt}],
                        max_tokens=350,
                        temperature=0,
                    )
                    intent_text = (intent_resp.choices[0].message.content or "").strip()
                    logger.info("Intent LLM response: %s", intent_text)
                    if "{" in intent_text and "}" in intent_text:
                        import json as std_json
                        raw_json = intent_text[intent_text.find("{"):intent_text.rfind("}")+1]
                        query_plan = std_json.loads(raw_json)
                        data_result = run_data_query(file_hash, query_plan.get("type"), query_plan.get("params", {}))
                        logger.info("Data Agent result: %s", str(data_result)[:500])

                        # ── Fallback: if data_result is an error or empty, try a broad search ──
                        result_is_empty = (
                            not data_result
                            or "error" in data_result
                            or (isinstance(data_result.get("result"), list) and len(data_result["result"]) == 0)
                            or data_result.get("result") == "No rows found."
                        )
                        if result_is_empty and resolved_subject:
                            logger.info("Primary query empty/failed — fallback search for '%s'", resolved_subject)
                            data_result = run_data_query(file_hash, "search", {"value": resolved_subject, "n": 3})

                    else:
                        # Intent LLM returned NONE — apply heuristic fallback for lookup-type questions
                        _LOOKUP_SIGNALS = {
                            "birthday", "born", "dob", "date of birth", "birth date",
                            "age", "address", "nationality", "country", "team", "club",
                            "salary", "height", "weight", "role", "position", "rank",
                            "when", "where", "who", "what is his", "what is her",
                            "what is their", "tell me about", "details of", "info on",
                        }
                        q_lower_check = question.lower()
                        is_lookup_question = any(sig in q_lower_check for sig in _LOOKUP_SIGNALS)
                        # Also trigger if the question has a pronoun and we have a resolved subject
                        if (is_lookup_question or has_pronoun) and resolved_subject:
                            logger.info("Intent=NONE but looks like a lookup — searching for '%s'", resolved_subject)
                            data_result = run_data_query(file_hash, "search", {"value": resolved_subject, "n": 3})
                        elif is_lookup_question:
                            # Try to extract a capitalized entity directly from the question
                            fallback_entity = _extract_recent_entity([{"content": question, "role": "user"}])
                            if fallback_entity:
                                logger.info("Intent=NONE fallback entity search: '%s'", fallback_entity)
                                data_result = run_data_query(file_hash, "search", {"value": fallback_entity, "n": 3})

                except Exception as e:
                    logger.warning("Intent pass failed (skipping data retrieval): %s", e)

            # ── Phase 2: Synthesis ───────────────────────────────────────────
            if data_result:
                # Inject resolved-subject hint so the synthesis LLM knows who we searched for
                subject_note = f" (subject resolved as: {resolved_subject})" if resolved_subject else ""
                messages.append({
                    "role": "system",
                    "content": (
                        f"ADDITIONAL DATA FROM FULL DATASET QUERY{subject_note}:\n{data_result}\n"
                        "IMPORTANT: The above is live data from the actual dataset. "
                        "For 'search' results, each item in 'result' is a full row — read the relevant field and state the value directly. "
                        "Do NOT say 'I don't have that data'. Parse the row and answer."
                    )
                })

            synthesis_model = os.getenv("GROQ_SYNTHESIS_MODEL", os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile"))
            completion = client.chat.completions.create(
                model=synthesis_model,
                messages=messages,
                temperature=0.15,
                max_tokens=450,
            )
            answer = (completion.choices[0].message.content or "").strip()
            if answer:
                return {"answer": answer, "data_queried": bool(data_result)}
        except Exception as exc:
            logger.warning("Groq chat failed: %s", exc)


    q_lower = question.lower()
    row_count = stats.get("row_count", "unknown")
    col_count = stats.get("column_count", "unknown")
    completeness = quality.get("completeness")
    missing_cells = quality.get("missing_cells")
    duplicate_rows = quality.get("duplicate_rows")
    findings = insights.get("findings") or []

    if "outlier" in q_lower:
        if top_outliers and top_outliers[0][1] > 0:
            summary = ", ".join([f"{col}: {cnt}" for col, cnt in top_outliers[:5]])
            return {"answer": f"Top outlier columns in {file_name} are {summary}."}
        return {"answer": f"No outlier counts are present for {file_name}, or all detected counts are zero."}

    if any(k in q_lower for k in ["quality", "missing", "duplicate", "completeness"]):
        parts = [f"For {file_name}, data quality shows {missing_cells or 0} missing cells and {duplicate_rows or 0} duplicate rows."]
        if completeness is not None:
            parts.append(f"Completeness is {float(completeness):.2f}%.")
        return {"answer": " ".join(parts)}

    if "correlation" in q_lower:
        if correlations:
            top = correlations[0]
            return {"answer": f"The strongest reported correlation is {top.get('col1')} and {top.get('col2')} with r={float(top.get('correlation', 0)):.3f}."}
        return {"answer": "No strong correlations were provided in the current analysis context."}

    # ── Greeting / non-data fallback — never dump stats on a casual message ──
    is_greeting_fallback = any(g in q_lower for g in ["hello", "hi", "hey", "thanks", "thank you", "bye", "okay", "ok"])
    if is_greeting_fallback:
        return {"answer": "Hello! I'm your data analyst. Feel free to ask me anything about your dataset."}

    # Generic data summary (only reached for genuinely stat-related questions)
    parts = [f"I analyzed {file_name} with {row_count} rows and {col_count} columns."]
    if completeness is not None:
        parts.append(f"Data completeness is {float(completeness):.2f}%.")
    if findings:
        parts.append(f"Key finding: {findings[0]}")
    return {"answer": " ".join(parts)}